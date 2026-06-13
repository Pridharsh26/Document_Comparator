import io
import logging
from typing import List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from models.schemas import ComparisonResult

logger = logging.getLogger(__name__)


class ExcelExportError(Exception):
    pass


def export_excel(comparison: ComparisonResult) -> bytes:
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparison Summary"

        summary_data = [
            ["Metric", "Value"],
            ["Total Legacy Rules", comparison.summary.total_legacy_rules],
            ["Total Modernized Rules", comparison.summary.total_modernized_rules],
            ["Added", comparison.summary.added],
            ["Removed", comparison.summary.removed],
            ["Modified", comparison.summary.modified],
            ["Unchanged", comparison.summary.unchanged],
        ]
        for row in summary_data:
            ws.append(row)

        changes_ws = wb.create_sheet(title="Changes")
        rows = []
        for change in comparison.changes:
            rows.append({
                "Change Type": change.change_type.value,
                "Legacy Rule": change.legacy_rule,
                "Modernized Rule": change.modernized_rule,
                "Confidence": change.matched_confidence,
                "Impact Level": change.impact_level,
                "Risk Level": change.risk_level,
                "Priority Level": change.priority_level,
                "Explanation": change.business_explanation,
            })
        df = pd.DataFrame(rows)
        for r in dataframe_to_rows(df, index=False, header=True):
            changes_ws.append(r)

        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()
    except Exception as exc:
        logger.error("Excel export failed: %s", exc)
        raise ExcelExportError(str(exc))


import io